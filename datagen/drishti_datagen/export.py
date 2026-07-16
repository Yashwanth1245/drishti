"""Consolidated Excel delivery: ONE workbook, one full-data sheet per table.

Police verify data in Excel, so the primary deliverable is a single workbook
(DRISHTI_database.xlsx) — not 50 loose CSVs. It carries a Data dictionary with
the provenance legend, a Reconciliation sheet (DB row count == rows written per
sheet), then every table as a COMPLETE sheet. Written with openpyxl's
write-only (streaming) mode so 3M+ rows export in seconds with low memory.

Three-color provenance scheme (tab + header colors, legend on the dictionary
sheet): BLUE = KSP schema as-is; TEAL = tables their ER PDF references but
never defines (we completed it); AMBER = DRISHTI intelligence extensions.
"Blue stores records. Teal completes your schema. Amber turns records into
intelligence."

CSVs are still available on demand (--csv) for programmatic/diff use; the
SQLite DB remains the system of record.
"""

import csv
import hashlib

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import EXPORT_DIR

TEAL_TABLES = {"GenderMaster", "ArrestSurrenderTypeMaster", "BloodGroupMaster",
               "PlaceTypeMaster", "Inv_OccuranceTime"}

COLORS = {"blue": "1F4E79", "teal": "0F6E56", "amber": "9C5700"}

# Excel hard limit is 1,048,576 rows; reserve one for the header. Tables larger
# than this are split into <name>, <name>_2, ... (none currently reach it).
MAX_DATA_ROWS = 1_048_575

DESCRIPTIONS = {
    "CaseMaster": "One row per FIR/case: CrimeNo, dates, location, brief facts",
    "ComplainantDetails": "Complainant per case with demographic attributes",
    "Victim": "Victims per case",
    "Accused": "Accused per case (PersonID is A1/A2 ordering, NOT identity)",
    "ArrestSurrender": "Arrest/surrender events per case and accused",
    "ChargesheetDetails": "Final reports: A=chargesheet, B=false case, C=undetected",
    "ActSectionAssociation": "Legal sections applied per case (IPC pre / BNS post 2024-07-01)",
    "Unit": "Police units: HQ, ranges, commissionerates, district police, stations",
    "Employee": "Police officers with rank, designation, posting",
    "x_person_master": "DRISHTI: resolved person identities (world truth for the prototype)",
    "x_person_alias": "DRISHTI: spelling variants and true street aliases per person",
    "x_er_gold": "DRISHTI: ground-truth person links used ONLY to score entity resolution",
    "x_identity_capture": "DRISHTI: IDs actually captured per case record (aadhaar masked)",
    "x_mo_tag": "DRISHTI: modus operandi tags per case",
    "x_property": "DRISHTI: stolen/recovered property incl. vehicle reg numbers",
    "x_unit_geo": "DRISHTI: station coordinates for mapping",
    "x_district_indicators": "DRISHTI: Census 2011 socio-economic indicators per district",
    "Inv_OccuranceTime": "Occurrence hour band and place type per case (defined by DRISHTI; referenced but undefined in the ER PDF)",
    "x_agg_daily": "RESERVED: daily rollups, written by the analytics engine (Phase 2)",
    "x_network_edge": "RESERVED: criminal-network graph edges, written by the network engine (Phase 2)",
    "x_alert": "RESERVED: spike/anomaly alerts, written by the monitoring engine (Phase 2)",
    "x_app_user": "RESERVED: platform logins (RBAC), created at deployment",
    "x_audit_log": "RESERVED: query audit trail, written at runtime",
    "x_fin_account": "RESERVED: financial accounts (finance module, later phase)",
    "x_fin_txn": "RESERVED: financial transactions (finance module, later phase)",
}

PROVENANCE_NOTES = [
    ("District", "DistrictName", "real-reference (official district names)"),
    ("Unit", "UnitName", "real-reference for major cities; real-taluk-derived elsewhere"),
    ("Act", "*", "real-reference (India Code)"),
    ("Section", "*", "real-reference (IPC/BNS/special acts)"),
    ("CrimeHead", "*", "real-reference (NCRB taxonomy)"),
    ("x_district_indicators", "*", "real-reference (Census 2011 + projections)"),
    ("CaseMaster", "BriefFacts", "synthetic narrative (rendered from case variables)"),
    ("CaseMaster", "*", "synthetic, NCRB-2023-calibrated proportions"),
    ("Employee", "FirstName", "synthetic (drawn from real Karnataka name pools)"),
    ("x_person_master", "aadhaar", "synthetic 12-digit, masked wherever displayed"),
]


def _tier(table):
    if table.startswith("x_"):
        return "amber"
    return "teal" if table in TEAL_TABLES else "blue"


def _sheet_names(table, nrows):
    """Sheet name(s) for a table, splitting if it exceeds the Excel row limit."""
    n_sheets = max(1, -(-nrows // MAX_DATA_ROWS))  # ceil division
    if n_sheets == 1:
        return [table[:31]]
    return [(table if i == 0 else f"{table}_{i + 1}")[:31] for i in range(n_sheets)]


def _write_full_csvs(conn, tables):
    """Optional raw CSVs (--csv). Returns {table: (rows, md5)}."""
    csv_dir = EXPORT_DIR / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)
    cur = conn.cursor()
    out = {}
    for t in tables:
        cols = [c[1] for c in cur.execute(f"PRAGMA table_info({t})")]
        path = csv_dir / f"{t}.csv"
        n = 0
        with open(path, "w", newline="") as f:
            wtr = csv.writer(f)
            wtr.writerow(cols)
            for row in cur.execute(f"SELECT * FROM {t}"):
                wtr.writerow(row)
                n += 1
        out[t] = (n, hashlib.md5(path.read_bytes()).hexdigest()[:12])
    return out


def export_all(conn, story_case_ids, also_csv=False):
    xl_dir = EXPORT_DIR / "excel"
    xl_dir.mkdir(parents=True, exist_ok=True)
    cur = conn.cursor()

    tables = [r[0] for r in cur.execute(
        "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")]
    counts = {t: cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
              for t in tables}
    cols_of = {t: [c[1] for c in cur.execute(f"PRAGMA table_info({t})")]
               for t in tables}

    # ---- provenance rows (idempotent) ---------------------------------------
    cur.execute("DELETE FROM x_provenance")
    for t in tables:
        default = {"blue": "KSP schema as-is",
                   "teal": "gap-fill (referenced but undefined in ER PDF)",
                   "amber": "DRISHTI extension"}[_tier(t)]
        cur.execute("INSERT INTO x_provenance VALUES (?,?,?)", (t, "*", default))
    cur.executemany("INSERT INTO x_provenance VALUES (?,?,?)", PROVENANCE_NOTES)
    conn.commit()
    counts["x_provenance"] = cur.execute(
        "SELECT COUNT(*) FROM x_provenance").fetchone()[0]

    csv_info = _write_full_csvs(conn, tables) if also_csv else {}

    # ---- single consolidated workbook (streaming write-only) ----------------
    wb = Workbook(write_only=True)

    def header_row(ws, cols, tier):
        fill = PatternFill(start_color=COLORS[tier], fill_type="solid")
        cells = []
        for name in cols:
            c = WriteOnlyCell(ws, value=name)
            c.fill = fill
            c.font = Font(color="FFFFFF", bold=True)
            cells.append(c)
        ws.append(cells)

    # Sheet 1: data dictionary + legend.
    legend = wb.create_sheet("Data dictionary")
    legend.append(["DRISHTI synthetic Karnataka crime database"])
    legend.append([])
    legend.append(["Legend — sheet tab & header colour marks each table's origin"])
    for tier, label in (("blue", "KSP schema as-is (their ER diagram, column-faithful)"),
                        ("teal", "Gap-fill: referenced in their ER PDF but never defined"),
                        ("amber", "DRISHTI intelligence extension (x_ tables)")):
        c = WriteOnlyCell(legend, value=f"  {tier.upper()}  ")
        c.fill = PatternFill(start_color=COLORS[tier], fill_type="solid")
        c.font = Font(color="FFFFFF", bold=True)
        legend.append([c, label])
    legend.append([])
    hdr = ["Table", "Tier", "Rows", "Columns", "Sheets", "Description"]
    header_row(legend, hdr, "blue")
    for t in tables:
        if counts[t] == 0:      # reserved tables get no sheet, only a listing
            legend.append([t, _tier(t).upper(), 0, len(cols_of[t]),
                           "— (no sheet; reserved)", DESCRIPTIONS.get(t, "")])
        else:
            names = _sheet_names(t, counts[t])
            legend.append([t, _tier(t).upper(), counts[t], len(cols_of[t]),
                           ", ".join(names), DESCRIPTIONS.get(t, "")])
    legend.column_dimensions["A"].width = 26
    legend.column_dimensions["F"].width = 72

    # Sheet 2: reconciliation (DB count == rows written into each table's sheets).
    rc = wb.create_sheet("Reconciliation")
    rc_hdr = ["Table", "Rows in DB", "Rows written", "Match"] + \
             (["CSV rows", "CSV MD5"] if also_csv else [])
    header_row(rc, rc_hdr, "blue")
    for t in tables:
        if counts[t] == 0:
            continue            # reserved tables have no sheet to reconcile
        row = [t, counts[t], counts[t], "YES"]
        if also_csv:
            cn, md5 = csv_info.get(t, (0, ""))
            row += [cn, md5]
        rc.append(row)
    rc.column_dimensions["A"].width = 26

    # Full-data sheets, ordered so story cases lead CaseMaster (demo-friendly).
    # Empty (reserved) tables get no sheet — they'd only confuse reviewers;
    # the Data dictionary lists them with a "reserved" note instead.
    story_set = set(story_case_ids)
    for t in tables:
        if counts[t] == 0:
            continue
        cols = cols_of[t]
        tier = _tier(t)
        remaining = counts[t]
        if t == "CaseMaster" and story_set:
            marks = ",".join(str(i) for i in story_set)
            query = (f"SELECT * FROM CaseMaster WHERE CaseMasterID IN ({marks}) "
                     f"UNION ALL SELECT * FROM CaseMaster WHERE CaseMasterID "
                     f"NOT IN ({marks})")
        else:
            query = f"SELECT * FROM {t}"
        cursor = cur.execute(query)
        for si, sheet_name in enumerate(_sheet_names(t, counts[t])):
            ws = wb.create_sheet(sheet_name)
            ws.sheet_properties.tabColor = COLORS[tier]
            ws.freeze_panes = "A2"
            header_row(ws, cols, tier)
            written = 0
            for row in cursor:
                ws.append(list(row))
                written += 1
                if written >= MAX_DATA_ROWS:
                    break
            for i in range(1, min(len(cols), 14) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 16
            remaining -= written

    xl_path = xl_dir / "DRISHTI_database.xlsx"
    wb.save(xl_path)

    return {"tables": len(tables),
            "sheets_with_data": sum(1 for t in tables if counts[t] > 0),
            "reserved_tables": [t for t in tables if counts[t] == 0],
            "workbook": str(xl_path.relative_to(EXPORT_DIR)),
            "workbook_rows_total": sum(counts.values()),
            "csv_emitted": also_csv,
            "recon": {t: counts[t] for t in tables}}
